from pathlib import Path

import requests
import json
import re
from datetime import datetime
from openpyxl import load_workbook
import hashlib
from timetable.stringlistanalyzer import StringListAnalyzer
from timetable.models import Resource, FileVersion


class FileData:
    """
    Неизменяемый класс, который хранит все параметры нового файла
    """
    TEMP_DIR = r"C:\Users\Ilya\Documents\Файлы сервера\temp"

    # Доверительное значение для определения параметров файла
    _CONFIDENCE_VALUE = 0.8

    # Слова, которые встречаются в степени обучения
    _DEGREE_WORDS = [
        "бакалавриат",
        "специалитет",
        "магистратура",
        "аспирантура",
        "степень"
    ]

    # Слова, которые встречаются в форме обучения
    _EDUCATION_FORM_WORDS = [
        "форма",
        "очная",
        "очно-заочная",
        "заочная"
    ]

    # Слова, которые встречаются в наименовании факультета
    _FACULTY_WORDS = [
        "факультет",
        "автоматизированных",
        "систем",
        "транспорта",
        "вооружений",
        "автомобильного",
        "технологии",
        "конструкционных",
        "материалов",
        "пищевых",
        "производств",
        "экономика",
        "управление",
        "электроника",
        "вычислительная",
        "техника",
        "xимико-технологический",
        "иностранный",
        "вечерний",
        "технологический",
        "инженерный",
        "кадры"
    ]

    # Слова, которые могут ошибочно появится в наименовании файла
    _WORDS_TO_DELETE = [
        "автосохраненный",
        "копия"
    ]

    # Слова, которые являются маркерами курса обучения
    _COURSE_WORDS = [
        "курс",
        "год"
    ]

    # Символы, которые могут разделять наименования
    _SENTENCE_DELIMITERS = ["_", " ", "(", ")", ",", ".", '"']

    # Расширения экселевских таблицы
    _EXCEL_EXTENSION = [".xls", ".xlsx"]

    def __init__(self, path:str, url:str, last_update):
        """
        Расчёт всех параметров для файла с параметрами
        :param path: Путь к файлу (содержит имя файла, предложенное в иерархии)
        :param url: Ссылка для скачивания
        :param last_update: Последнее время обновления
        """

        self.__path = path # Путь по которому хранился файл на сайте
        self.__url = url # Ссылка для скачивания файла
        self.__last_changed = last_update # Время последнего обновления
        self.__calc()

    def get_path(self) -> str: return self.__path
    def get_url(self) -> str: return self.__url
    def get_last_changed(self) -> str: return self.__last_changed

    def __calc(self):
        """
        Рассчитывает все параметры для файла
        :return: Текущий экземпляр класса
        """
        self.__name_from_path_with_mimetype = self.get_file_name_from_path(self.__path, dell_mimetype=False)  # Имя файла с расширением из пути
        self.__name_from_path = self.get_file_name_from_path(self.__path)  # Имя файла с расширением из пути
        self.__mimetype_from_path = self.__get_mimetype(self.__path) # Расширение файла из пути
        self.__correct_name_from_path = self.get_correct_file_name(self.__name_from_path)  # Корректное имя файла из пути

        self.__name_from_url_with_mimetype = self.get_file_name_from_path(self.__url, dell_mimetype=False) # Имя файла с расширением из ссылки
        self.__name_from_url = self.get_file_name_from_path(self.__url) # Имя файла из ссылки
        self.__mimetype_from_url = self.__get_mimetype(self.__path) # Расширение файла из ссылки
        self.__correct_name_from_url = self.get_correct_file_name(self.__correct_name_from_path) # Корректное имя файла из url

        self.__degree = self._get_degree(self.__path)  # Степень обучения (бакалавриат, магистратура и другие)
        self.__education_form = self._get_education_form(self.__path)  # Форма обучения (очная, заочная и другие)
        self.__faculty = self._get_faculty(self.__path)  # Факультет обучения
        self.__course = self._get_course_list(self.__name_from_path) # Список курсов

        self.__correct_path = self.__calc_correct_path(self.__path)  # Корректный путь к файлу
        self.__hash = None

        return self

    def get_name(self): return self.__name_from_path
    def get_mimetype(self): return self.__mimetype_from_url
    def get_file_name(self): return self.__name_from_url_with_mimetype

    def get_degree(self) -> str: return self.__degree
    def get_education_form(self) -> str: return self.__education_form
    def get_faculty(self) -> str: return self.__faculty
    def get_course(self) -> str: return str(self.__course)

    def get_name_from_path(self) -> str: return self.__name_from_path
    def get_correct_name_from_path(self) -> str: return self.__correct_name_from_path
    def get_name_from_url(self) -> str: return self.__name_from_url
    def get_correct_name_from_url(self) -> str: return self.__correct_name_from_url
    def get_correct_path(self) -> str: return self.__correct_path

    @classmethod
    def _get_degree(cls, path: str | list):
        """
        Ищет степень обучения в пути файла.
        Путь может быть представлен в виде списка директорий, так и классической строки "dir1/dir2/dir3/".
        :param path: Путь файла
        :return: Степень обучения
        """
        if isinstance(path, str):
            path = path.split('/')
        return cls.__get_best_element(path, lambda string: cls.__get_degree_word_count(string))

    @classmethod
    def _get_education_form(cls, path: str | list):
        """
        Ищет форму обучения в пути файла.
        Путь может быть представлен в виде списка директорий, так и классической строки "dir1/dir2/dir3/".
        :param path: Путь файла
        :return: Форма обучения
        """
        if isinstance(path, str):
            path = path.split('/')
        return cls.__get_best_element(path, lambda string: cls.__get_education_form_word_count(string))

    @classmethod
    def _get_faculty(cls, path: str | list):
        """
        Ищет факультет в пути файла.
        Путь может быть представлен в виде списка директорий, так и классической строки "dir1/dir2/dir3/"
        :param path: Путь файла
        :return: Факультет
        """
        if isinstance(path, str):
            path = path.split('/')
        return cls.__get_best_element(path, lambda string: cls.__get_faculty_word_count(string))

    @classmethod
    def __get_course_string(cls, courses:list[int]):
        """
        Получить строку с номером курса
        :param courses:
        :return:
        """
        if len(courses) == 0:
            return "Неопределенно"
        else:
            string = "Курс"
            for i in range(len(courses)):
                string += " " + str(courses[i])
                if i != len(courses) - 1:
                    string += ","
            return string

    @classmethod
    def _get_course_list(cls, string:str):
        """
        Найти номера курсов
        :param string: Строка с номером курса
        :return: Список номеров курсов
        """
        # Список чисел
        numbers = []

        # Поиск чисел по ключевым словам, пока они не будут найдены
        for mark_word in cls._COURSE_WORDS:
            # Расчёт чисел
            numbers = cls.__FindNumbersByMarkWord(string, mark_word).get_list()
            numbers = list(set(s for s in numbers if 0 < s < 7))

            # Завершить поиск, если есть числа
            if len(numbers) > 0:
                break

        # Вернуть список чисел
        return numbers

    class __FindNumbersByMarkWord:
        """
        Ищет в строке список чисел по ключевому слову.
        """
        def __init__(self, string: str, mark_word):
            """
            Ищет в строке список чисел, которые соответствуют ключевому слову.
            :param string:
            :param mark_word:
            """
            self.result_list = self.__get_numbers_list_by_mark_word(string, mark_word)

        def get_list(self):
            """
            Возвращает список чисел.
            :return: Список чисел
            """
            return self.result_list

        @classmethod
        def __get_numbers_list_by_mark_word(cls, string: str, mark_word:str):
            """
            Возвращает список чисел, которые соответствуют ключевому слову.
            Если ни одно слово не похоже на ключевое, список будет пуст.
            :param string: Строка, в которой содержатся числа и ключевое слово
            :param mark_word: Ключевое слово
            :return: Список чисел
            """
            # Список чисел
            number_list = []

            # Поиск слов, максимально похожих на наше ключевое слово
            analyze = StringListAnalyzer(FileData.split_string_by_delimiters(string), [mark_word])
            mark_words_in_string = analyze.get_max_ratio_words()

            # Вернуть пустой список, если слово недостаточно похоже или ни одно слово не найдено
            mxs = analyze.get_max_ratio()
            if analyze.get_max_ratio() < FileData._CONFIDENCE_VALUE or len(mark_words_in_string) == 0:
                return []

            # Первое, максимально похожее слово
            mark_word_in_string = mark_words_in_string[0]
            # Длина максимально похожего слова
            mark_word_in_string_length = len(mark_word_in_string)

            # Поиск первого списка чисел
            for start_index in cls.__find_substring_occurrences(string, mark_word_in_string):
                # Конечный индекс слова
                finish_index = start_index + mark_word_in_string_length

                # Список слов слева и справа
                left_words = cls.__find_all_words(string[:start_index])[::-1]
                right_words = cls.__find_all_words(string[finish_index+1:])

                # Список строк чисел и диапазонов чисел слева и справа
                left_numbers = cls.__get_first_elements_in_list(left_words, lambda word: cls.__is_number_or_number_range(word))
                right_numbers = cls.__get_first_elements_in_list(right_words, lambda word: cls.__is_number_or_number_range(word))

                # Перейти к следующему, если числа не были найдены
                if len(left_numbers) == 0 and len(right_numbers) == 0:
                    continue

                # Если числа есть только с одной стороны, то завершить анализ
                if len(left_numbers) > 0 and len(right_numbers) == 0:
                    number_list = cls.__get_number_list(left_numbers)
                    break
                elif len(left_numbers) == 0 and len(right_numbers) > 0:
                    number_list = cls.__get_number_list(right_numbers)
                    break
                else:
                    number_list = cls.__get_number_list(left_numbers + right_numbers)
                    break

            return number_list

        @staticmethod
        def __find_substring_occurrences(string:str, substring):
            """
            Найти все вхождения строки в подстроку и вернуть их индексы.
            :param string: Строка
            :param substring: Подстрока
            :return: Список индексов
            """
            # Список индексов вхождений
            indices = []

            # Поиск всех вхождений
            index = string.find(substring)
            while index != -1:
                # Запись индекса в список
                indices.append(index)
                # Найти следующее вхождение
                index = string.find(substring, index + 1)

            # Вернуть список индексов
            return indices

        @staticmethod
        def __find_all_words(string:str):
            """
            Ищет все слова и диапазоны чисел.
            :param string: Строка
            :return: Список строк
            """
            return re.findall(r'\b(\d+\s*-\s*\d+|\w+)\b', string)

        @staticmethod
        def __get_first_elements_in_list(elements:list, condition):
            """
            Возвращает список первых элементов в списке удовлетворяющих условию.
            :param elements: Список элементов
            :param condition: Условие (lambda функция)
            :return: Список элементов
            """
            first_elements = []
            for element in elements:
                if condition(element):
                    first_elements.append(element)
                else:
                    break
            return first_elements

        @staticmethod
        def __is_number_or_number_range(string:str):
            """
            Проверка строки на число или диапазон чисел.
            :param string: Строка
            :return: Результат проверки
            """
            val = re.search(r'\b(\d+\s*-\s*\d+|\d+)\b', string)
            return re.search(r'\b(\d+\s*-\s*\d+|\d+)\b', string) is not None

        @staticmethod
        def __get_number_list(string_list:list[str]) -> list[int]:
            """
            Получить список чисел из списка строк с числами и диапазонами.
            :param string_list: Список строк
            :return: Список чисел
            """
            # Список чисел
            numbers = []

            # Для каждой строки
            for string in string_list:
                # Добавить диапазон
                if string.find('-') > 0:
                    left_number, right_number = string.split('-')[0:2]
                    numbers += list(range(int(left_number), int(right_number) + 1))
                # Добавить число
                elif string.isdigit():
                    numbers.append(int(string))
                # Выбросить исключение, т.к. строка не соответствует ни одному из типов
                else:
                    raise Exception(f'String is not digit and is not range. String: "{string}".')

            # Вернуть список чисел
            return sorted(numbers)

    @classmethod
    def get_file_name_from_path(cls, path: str, dell_mimetype:bool = True):
        """
        Вычисляет имя файла из пути. Путь представлен в виде классической строки "dir1/dir2/dir3/filename".
        :param path: Путь к файлу
        :param dell_mimetype: Удалить расширение из имени
        :return: Имя файла
        """
        # Получить имя файла с его расширением
        file_name = path.split('/')[-1]

        # Удалить расширение файла, если это необходимо
        if dell_mimetype:
            file_name = re.sub(r'\.[А-ЯЁA-Zа-яёa-z]*$', "", file_name)

        # Вернуть имя файла
        return file_name

    @classmethod
    def __get_mimetype(cls, name:str) -> str:
        """
        Получает mimetype из имени файла.
        :param name: Имя файла
        :return: mimetype
        """
        # Вернуть имя файла
        return name.split('.')[-1]

    @classmethod
    def get_correct_file_name(cls, file_name):
        """
        Вернуть корректное имя файла. Удаляет запрещённые слова и лишние символы.
        :param file_name: Имя файла
        :return: Новое имя файла
        """
        # Удаление лишние пробелы
        new_file_name = cls.__remove_extra_spaces(file_name)

        # Поиск слов, максимально похожих на список запрещённых
        analyze = StringListAnalyzer(cls.split_string_by_delimiters(new_file_name), cls._WORDS_TO_DELETE)
        del_words = analyze.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1)
        # Удаление запрещённых слов
        for word in del_words:
            new_file_name = new_file_name.replace(word, "")

        # Удаление бесполезных строковых конструкций
        while True:
            # Запомнить версию строки до преобразований
            before_convert_string = new_file_name

            # Преобразование над строкой
            new_file_name = re.sub(r'-\s*$', '', new_file_name)
            new_file_name = re.sub(r'\(\s*\)', "", new_file_name)
            new_file_name = cls.__remove_extra_spaces(new_file_name)

            # Выход из цикла, если строка не изменилась
            if new_file_name == before_convert_string:
                break

        # Вернуть преобразованную строку
        return new_file_name

    def __calc_correct_path(self, path:str = "", number_of_first_directories:int = 0):
        """
        Вычисляет корректный путь. Может добавить несколько первых директорий из заданного пути в корректный.
        Путь представлен в виде классической строки "dir1/dir2/dir3/filename".
        Путь строится по принципу: "dir1/.../dirN/{Степень обучения}/{Факультет}/{Форма обучения}/{Курс}/".
        :param path: Исходный путь
        :param number_of_first_directories: Количество первых директорий, которые будут включены в конечный путь
        :return: Корректный путь
        """
        # Добавить несколько первых директорий в новый путь, если это необходимо
        dirs = path.split('/')
        new_path = dirs[:number_of_first_directories]

        new_path.append(self.__degree) # Добавить степень обучения
        new_path.append(self.__faculty) # Добавить факультет
        new_path.append(self.__education_form) # Добавить форму обучения
        new_path.append(self.__get_course_string(self.__course)) # Добавить курс

        # Вернуть новый путь
        return self.elements_to_path(new_path)

    @classmethod
    def elements_to_path(cls, elements: list, base_path:str = "", is_file=False):
        """
        Добавляет к уже существующему пути дополнительные пути
        :param elements: Список элементов
        :param base_path: Изначальный путь
        :param is_file: Последний элемент является файлом, символ '/' в конце строки не нужен
        :return: Дополненный путь
        """
        result_path = base_path
        list_length = len(elements)
        for i in range(list_length):
            result_path = cls.add_to_path(result_path, elements[i], i + 1 >= list_length and is_file)
        return result_path

    @staticmethod
    def add_to_path(path, element: str, is_file=False):
        """
        Добавляет элемент к пути. Возвращает текущую строку, если элемент отсутствует.
        :param path: Изначальный путь
        :param element: Следующий элемент в пути
        :param is_file: Элемент является файлом, символ '/' в конце строки не нужен
        :return: Дополненный путь
        """
        if element != "" and element is not None:
            path += element
            if not is_file:
                path += '/'
        return path

    @staticmethod
    def __remove_extra_spaces(string:str):
        return re.sub(r'\s+', ' ', string).strip()

    @staticmethod
    def __get_best_element(elements:list, heuristic_function, min_heuristic = 0):
        """
        Определяет наилучший элемент в списке по эвристической функции
        :param elements: Список элементов
        :param heuristic_function: Эвристическая функция (lambda функция)
        :param min_heuristic: Минимально возможное значение эвристики
        :return: Наилучший элемент
        """
        # Максимальное значение эвристики
        max_heuristic = min_heuristic

        # Элемент с наилучшей эвристикой
        best_element = None

        # Поиск наилучшего значения
        for element in elements:
            heuristic = heuristic_function(element)
            if heuristic > max_heuristic:
                max_heuristic = heuristic
                best_element = element

        # Вернуть наилучший элемент
        return best_element

    @classmethod
    def __get_degree_word_count(cls, string:str):
        """
        Расчёт эвристического значения для поиска степени обучения.
        Вычисляется количество слов в строке, которые максимально похожи
        на слова из списка слов-маркеров для степени обучения.
        :param string: Строка для поиска
        :return: Эвристическое значение
        """
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._DEGREE_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))

    @classmethod
    def __get_education_form_word_count(cls, string:str):
        """
        Расчёт эвристического значения для поиска формы обучения.
        Вычисляется количество слов в строке, которые максимально похожи
        на слова из списка слов-маркеров для формы обучения.
        :param string: Строка для поиска
        :return: Эвристическое значение
        """
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._EDUCATION_FORM_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))

    @classmethod
    def __get_faculty_word_count(cls, string:str):
        """
        Расчёт эвристического значения для поиска факультета.
        Вычисляется количество слов в строке, которые максимально похожи
        на слова из списка слов-маркеров для факультета.
        :param string: Строка для поиска
        :return: Эвристическое значение
        """
        analyzer = StringListAnalyzer(cls.split_string_by_delimiters(string.lower()), cls._FACULTY_WORDS)
        return len(analyzer.get_strings_by_ratio_in_range(cls._CONFIDENCE_VALUE, 1))

    @classmethod
    def split_string_by_delimiters(cls, string:str, delimiters:list|None=None) -> list[str]:
        """
        Разделить строку на слова по разделителям
        :param string: Строка
        :param delimiters: Список разделителей
        :return: Список слов
        """
        # Используется стандартный набор разделителей если другой не указан явно
        if delimiters is None:
            delimiters = cls._SENTENCE_DELIMITERS

        # Создаем регулярное выражение для поиска всех разделителей
        regex_pattern = '|'.join(map(re.escape, delimiters))

        # Разделяем строку по заданному шаблону
        return re.split(regex_pattern, string)

    def get_json(self):
        """
        Создаёт json файл с параметрами ресурса
        :return:
        """
        json_data = {
            "degree": self.get_degree(),
            "education_form": self.get_education_form(),
            "faculty": self.get_faculty(),
            "course": self.get_course()
        }
        return json.dumps(json_data, indent=4)

    def get_resource(self):
        """
        Возвращает новую запись ресурса для базы данных.
        :return: Модель записи базы данных
        """
        new_resource = Resource()
        new_resource.path = self.get_correct_path()
        new_resource.name = self.get_name()
        new_resource.metadata = self.get_json()
        return new_resource

    def get_file_version(self, file_path:Path|str, resource_id:int = None):
        """
        Возвращает новую запись версии файла для базы данных.
        :param file_path: Путь к файлу, для которой создаётся запись
        :param resource_id: ID ресурса, к которому привязан файл
        :return: Модель записи базы данных
        """
        # Приводим тип к нужному типу
        file_path = Path(file_path)
        if not file_path.is_file():
            raise FileNotFoundError

        # Создаём новый объект базы данных
        new_file_version = FileVersion()

        #Добавить ID ресурса, если он есть
        if resource_id is not None:
            new_file_version.resource = resource_id

        # Добавить остальные параметры
        new_file_version.mimetype = file_path.suffix # Добавить расширение файла
        # Добавить время обновления, указанное на сайте
        try:
            new_file_version.last_changed = datetime.strptime(self.get_last_changed(), "%Y-%m-%d %H:%M:%S")
        except ValueError:
            # Вернуть текущее время, если не получилось распарсить время с сайта
            new_file_version.last_changed = datetime.now()
        new_file_version.url = self.get_url() # Добавить ссылку на скачивание
        new_file_version.hashsum = self.__get_file_hash(file_path) # Добавляет hash сумму файла

        # Вернуть запись
        return new_file_version

    @classmethod
    def __get_file_hash(cls, file_path:Path):
        suffix = file_path.suffix
        if suffix in cls._EXCEL_EXTENSION:
            print(suffix)
            return cls.__get_excel_file_hash(file_path)
        else:
            return cls.__get_bin_file_hash(file_path)

    @staticmethod
    def __get_excel_file_hash(file_path:Path):
        wb = load_workbook(str(file_path), data_only=True)
        data = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                data.append(tuple(row))  # Собираем данные
        return hashlib.sha256(str(data).encode('utf-8')).hexdigest()

    @staticmethod
    def __get_bin_file_hash(file_path:Path):
        # Инициализация SHA-256
        sha256_hash = hashlib.sha256()

        # Открываем файл в режиме чтения байтов
        with file_path.open("rb") as f:
            # Читаем файл поблочно для экономии памяти
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)

        # Возвращаем хэш файла в виде строки
        return sha256_hash.hexdigest()

    def download_file(self, dir: Path | str, chunk_size: int = 8192):
        """
        Скачивает файл по ссылке и сохраняет его по заданному пути
        :param dir: Путь для сохранения файла
        :param chunk_size: размер чанка для сохранения
        :return: Путь по которому сохранён файл
        """
        # Попытаться скачать файл
        download_file = requests.get(self.get_url())

        # Выбросить исключение если файл не был скачан успешно
        if download_file.status_code != 200:
            raise Exception(f"File download error. URL: {self.get_url()}")

        dir = Path(dir)
        # Создаём путь к папке
        dir.mkdir(parents=True, exist_ok=True)

        # Создать путь к файлу
        file_path = dir / self.get_file_name()

        # Сохранить файл
        with file_path.open('wb') as file:
            for chunk in download_file.iter_content(chunk_size=chunk_size):
                file.write(chunk)

        # Вернуть путь
        return file_path